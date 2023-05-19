import openpyxl, os, librosa

wb = openpyxl.load_workbook('data_automation\data.xlsx')
ex = wb[r'분석 데이터']

sounds = os.listdir('data_automation\sound_files')


for ro in range(2, 50):

    if ex.cell(row=ro, column=3).value != None:
        continue

    elif ex.cell(row=ro, column=1).value == None:
        break

    else:
        file_path = r'C:\Users\Jooney Han\Desktop\과학전람회\data\sound\15-길이-임계.wav'
        y, sr = librosa.load(file_path, sr=44100)  
        inst_time = ex.cell(row=ro, column=1).value
        print(inst_time)
        
        time_index = int(inst_time * sr)
        frame_size = 2048
        hop_length = 512

        # Calculate the STFT (Short-Time Fourier Transform)
        D = librosa.stft(y, n_fft=frame_size, hop_length=hop_length)

        magnitude = librosa.amplitude_to_db(abs(D))
        pitches = librosa.yin(y, fmin=75, fmax=500, frame_length=frame_size, hop_length=hop_length)

        # Find the amp and freq at the specified time index
        fund_amp = max(magnitude[:, time_index // hop_length])
        fund_freq = max(pitches)

        
        # ex.cell(row=ro, column=2).value = str(sr) --> parameter of image
        ex.cell(row=ro, column=3).value = fund_amp
        ex.cell(row=ro, column=4).value = fund_freq

wb.save('data_automation\saved.xlsx')